import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import pytz
from helpers import sheet_names, range_specs, setup_header, setup_footer


# Set the Streamlit page configuration
st.set_page_config(page_title="PIT Combiner", page_icon=":house:", layout="wide")

def get_current_time(timezone: str) -> str:
    """
    Get the current time in a specified timezone and return it as a formatted string.
    
    Args:
        timezone (str): A timezone string recognized by pytz.
    
    Returns:
        str: The current time formatted as 'YYYY-MM-DD_HH-MM-SS'.
    """
    return datetime.now(pytz.timezone(timezone)).strftime('%Y-%m-%d_%H-%M-%S')

timezone = "America/Chicago"
current_time = get_current_time(timezone)

def clean_data(workbook, sheet_names_map, terms_to_delete):
    """
    Clean data in the given workbook by removing rows containing any of the specified terms.
    This top-level function is defined but overshadowed by a similar function within load_and_sum.
    It's maintained for consistency but not used directly.

    Args:
        workbook: An openpyxl Workbook object.
        sheet_names_map (dict): A dictionary mapping identifiers to sheet names.
        terms_to_delete (list): A list of terms that if found in a row, that row should be deleted.
    """
    for sheet_name in sheet_names_map.values():
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_to_delete = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    if any(term in str(cell.value) for term in terms_to_delete):
                        rows_to_delete.append(row[0].row)
                        break
            for row_num in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_num)


def load_and_sum(hmis_stream, non_hmis_stream, range_specs, sheet_names_map):
    """
    Load two Excel workbooks (HMIS and Non-HMIS), clean their data, and then sum up values
    as defined by the range_specs. The results are written to a template workbook which retains
    the original formulas. The summed results are placed in specified target cells.

    Args:
        hmis_stream (UploadedFile): Stream of the HMIS Excel file uploaded by user.
        non_hmis_stream (UploadedFile): Stream of the Non-HMIS Excel file uploaded by user.
        range_specs (list): A list of tuple specifications that define what ranges to sum and where.
        sheet_names_map (dict): A dictionary mapping identifiers to sheet names.

    Returns:
        BytesIO: A BytesIO stream of the resulting combined Excel workbook.
    """
    # Load the input workbooks with data_only=True to get computed values
    hmis_wb = load_workbook(filename=BytesIO(hmis_stream.read()), data_only=True)
    non_hmis_wb = load_workbook(filename=BytesIO(non_hmis_stream.read()), data_only=True)
    
    # Load the template workbook (data_only=False to preserve formulas)
    with open("template.xlsx", "rb") as template_file:
        template_wb = load_workbook(filename=BytesIO(template_file.read()), data_only=False)

    # Terms that, if present, will cause a row to be deleted
    terms_to_delete = [
        "Client Doesn't Know", "Client Prefers Not to Answer", "Data Not Collected",
        "Client Doesn't Know/Prefers Not to Answer", "Missing Information",
        "Number of persons missing DoB", "Null"
    ]

    def internal_clean_data(workbook, sheet_names_map, terms_to_delete):
        """
        Internal clean_data function that avoids deleting instruction rows and ensures that
        only actual data rows containing specified terms are removed.
        """
        for s_name in sheet_names_map.values():
            if s_name not in workbook.sheetnames:
                continue
            sheet = workbook[s_name]
            rows_to_remove = []
            for row in sheet.iter_rows():
                # Skip instruction rows (e.g., rows that might contain HUD instructions)
                if row[0].value and isinstance(row[0].value, str) and "HUD does not allow" in row[0].value:
                    continue
                # Check if the row contains any of the forbidden terms
                if any(cell.value in terms_to_delete for cell in row):
                    rows_to_remove.append(row[0].row)
            # Delete rows from bottom to top to avoid index shifting issues
            for row_idx in reversed(rows_to_remove):
                sheet.delete_rows(row_idx)

    # Clean data in both HMIS and Non-HMIS workbooks
    internal_clean_data(hmis_wb, sheet_names_map, terms_to_delete)
    internal_clean_data(non_hmis_wb, sheet_names_map, terms_to_delete)

    # Process the specified ranges and sum data accordingly
    for source_ranges, target_sheet_name, target_col, target_start_row, target_end_row in range_specs:
        if target_sheet_name not in template_wb.sheetnames:
            continue
        
        output_sheet = template_wb[target_sheet_name]

        # For each row in the range, compute the sum from HMIS and Non-HMIS sheets
        for row_offset in range(target_end_row - target_start_row + 1):
            sum_value = 0
            for sheet_name_key, col_prefix, start_row, end_row in source_ranges:
                # Determine which workbook to read from
                # Note: 'HDX' presence in sheet_name_key determines which workbook is used.
                workbook = non_hmis_wb if 'HDX' in sheet_name_key else hmis_wb
                sheet = workbook[sheet_names_map[sheet_name_key]]

                row_index = start_row + row_offset
                if row_index > end_row:
                    continue

                cell_value = sheet[f"{col_prefix}{row_index}"].value

                # Attempt to convert string values to float
                if isinstance(cell_value, str):
                    try:
                        cell_value = float(cell_value.replace(',', ''))
                    except ValueError:
                        cell_value = 0

                if isinstance(cell_value, (int, float)):
                    sum_value += cell_value

            # Only overwrite if the template cell does not contain a formula
            cell_ref = f"{target_col}{target_start_row + row_offset}"
            if not isinstance(output_sheet[cell_ref].value, str):
                output_sheet[cell_ref].value = sum_value

    # Save the modified template to a BytesIO object
    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output


# Set up the Streamlit layout and interface
setup_header()

st.warning("Upload the HMIS and Non-HMIS Excel files")

# File uploaders
hmis_file = st.file_uploader("Upload HMIS Data (HUDX_230AD)", type=['xlsx'])
non_hmis_file = st.file_uploader("Upload Non-HMIS Data", type=['xlsx'])

# Process button
if st.button("Process Files"):
    if hmis_file and non_hmis_file:
        # Call the processing function
        output_file = load_and_sum(hmis_file, non_hmis_file, range_specs, sheet_names)

        # Convert the BytesIO stream to bytes for download
        output_bytes = output_file.getvalue()

        # Download button for the output Excel file
        st.download_button(
            label="Download Output File",
            data=output_bytes,
            file_name=f"combined_data_{current_time}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Please upload both HMIS and Non-HMIS files to proceed.")

setup_footer()
